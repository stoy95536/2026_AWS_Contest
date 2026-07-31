"""
Excel 資料載入器
負責讀取多工作表 Excel，辨識欄位、月份、銀行名稱、數值與單位。
使用 openpyxl 直接讀取原生結構，不將 Excel 轉為圖片。
"""

import os
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter


class ExcelLoader:
    """讀取 Excel 檔案並提取原生結構資訊。"""

    def __init__(self, file_path: str):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"找不到檔案: {file_path}")
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet_names = self.workbook.sheetnames

    def get_sheet_names(self) -> list[str]:
        """取得所有工作表名稱。"""
        return self.sheet_names

    def read_sheet_to_dataframe(self, sheet_name: str, header_row: int = 1) -> pd.DataFrame:
        """
        將指定工作表讀取為 pandas DataFrame。

        Args:
            sheet_name: 工作表名稱
            header_row: 標題行號 (1-indexed)

        Returns:
            pandas DataFrame
        """
        df = pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            header=header_row - 1,
            engine="openpyxl",
        )
        return df

    def get_cell_value(self, sheet_name: str, row: int, col: int) -> Any:
        """取得指定儲存格的值。"""
        ws = self.workbook[sheet_name]
        return ws.cell(row=row, column=col).value

    def get_cell_reference(self, sheet_name: str, row: int, col: int) -> str:
        """取得儲存格參照字串，如 'Sheet1!A1'。"""
        col_letter = get_column_letter(col)
        return f"{sheet_name}!{col_letter}{row}"

    def get_merged_cells(self, sheet_name: str) -> list[str]:
        """取得合併儲存格範圍清單。"""
        ws = self.workbook[sheet_name]
        return [str(merged) for merged in ws.merged_cells.ranges]

    def detect_header_row(self, sheet_name: str, max_scan: int = 10) -> int:
        """
        偵測標題行位置（啟發式：第一行全為文字的列）。

        Args:
            sheet_name: 工作表名稱
            max_scan: 最多掃描幾列

        Returns:
            標題行號 (1-indexed)
        """
        ws = self.workbook[sheet_name]
        for row_idx in range(1, max_scan + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            non_none = [v for v in values if v is not None]
            if non_none and all(isinstance(v, str) for v in non_none):
                return row_idx
        return 1

    def get_sheet_dimensions(self, sheet_name: str) -> dict:
        """取得工作表的行列範圍。"""
        ws = self.workbook[sheet_name]
        return {
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_column": ws.min_column,
            "max_column": ws.max_column,
        }

    def extract_raw_data(self, sheet_name: str) -> list[list[Any]]:
        """取得工作表所有原始資料（二維列表）。"""
        ws = self.workbook[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def close(self):
        """關閉工作簿。"""
        self.workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
