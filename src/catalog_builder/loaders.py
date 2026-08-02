"""
輸入檔載入層：讓 .xlsx / .xls / .csv 走同一條解析路徑。

原本只掃 `*.xlsx`，目錄裡的 `.xls` 與 `.csv` 會被**靜默略過且不留任何警告**——
實測 3 個檔案的目錄只看到 1 個。決賽若 11 份裡混了 3 份舊格式，系統會安靜地
只算 8 份，產出一份看起來完整、實則少了四分之一資料的分析。台灣政府開放資料
很常以 `.xls` 發布，這個風險不低。

`structure_detector` 只需要工作表提供四樣東西：`title`、`max_row`／`max_column`、
`cell(row, col).value`、`merged_cells.ranges`。因此這裡用轉接器把 xlrd 與 csv
包成同樣的介面，解析邏輯一行都不用改。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

import openpyxl

XLSX_SUFFIXES = {".xlsx", ".xlsm"}
XLS_SUFFIXES = {".xls"}
CSV_SUFFIXES = {".csv", ".txt"}
SUPPORTED_SUFFIXES = XLSX_SUFFIXES | XLS_SUFFIXES | CSV_SUFFIXES

SPREADSHEET_LIKE = SUPPORTED_SUFFIXES | {".ods", ".xlsb", ".numbers"}
"""看起來像資料檔的副檔名。

不在支援清單但屬於這一組的檔案要**大聲報出來**——使用者把 .ods 放進來卻
沒被處理，比直接報錯危險得多。"""

CSV_ENCODINGS = ("utf-8-sig", "cp950", "utf-8", "big5")
"""CSV 編碼嘗試順序。

政府開放資料的 CSV 常是 cp950／big5；utf-8-sig 放第一是因為 Excel 匯出的
UTF-8 CSV 幾乎都帶 BOM，用純 utf-8 讀會在第一格多出 \\ufeff。"""


class UnsupportedFormatError(ValueError):
    """檔案格式無法解析。訊息會直接呈現給使用者，不吞掉。"""


# --------------------------------------------------------------------------
# 掃描
# --------------------------------------------------------------------------

@dataclass
class InputScan:
    """目錄掃描結果。"""

    supported: list[Path] = field(default_factory=list)
    unsupported: list[Path] = field(default_factory=list)
    """看起來是資料檔、但目前解析不了的檔案。必須讓使用者看到。"""

    @property
    def warnings(self) -> list[str]:
        if not self.unsupported:
            return []
        names = "、".join(p.name for p in self.unsupported[:5])
        more = f" 等 {len(self.unsupported)} 個" if len(self.unsupported) > 5 else ""
        return [
            f"以下檔案格式不支援，**未納入分析**：{names}{more}。"
            f"目前支援 {sorted(SUPPORTED_SUFFIXES)}；"
            "請先轉存成 .xlsx 再重跑，否則分析結果會少掉這些資料。"
        ]


def scan_inputs(data_dir: str | Path) -> InputScan:
    """
    掃描目錄，分出「能解析」與「像資料檔但解析不了」。

    第二類刻意單獨列出：靜默略過會讓使用者以為全部都算進去了。
    """
    src = Path(data_dir)
    scan = InputScan()

    for path in sorted(src.iterdir() if src.is_dir() else []):
        if not path.is_file() or path.name.startswith("~$"):
            continue  # ~$ 開頭是 Excel 開啟中的暫存檔
        suffix = path.suffix.lower()
        if suffix in SUPPORTED_SUFFIXES:
            scan.supported.append(path)
        elif suffix in SPREADSHEET_LIKE:
            scan.unsupported.append(path)

    return scan


# --------------------------------------------------------------------------
# 轉接器：把 xlrd／csv 包成 openpyxl 的介面
# --------------------------------------------------------------------------

class _Cell:
    """只需要 `.value`，`structure_detector` 不用其他屬性。"""

    __slots__ = ("value",)

    def __init__(self, value: Any):
        self.value = value


class _MergedRange:
    """對齊 openpyxl 的 `CellRange`，提供 1-indexed 的四個邊界。"""

    __slots__ = ("min_row", "max_row", "min_col", "max_col")

    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int):
        self.min_row, self.max_row = min_row, max_row
        self.min_col, self.max_col = min_col, max_col

    def __str__(self) -> str:
        from openpyxl.utils import get_column_letter

        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


class _MergedCells:
    """openpyxl 的 `ws.merged_cells` 是物件而非 list，這裡對齊它的 `.ranges`。"""

    __slots__ = ("ranges",)

    def __init__(self, ranges: list[_MergedRange]):
        self.ranges = ranges


class GridWorksheet:
    """
    以二維陣列為底的工作表，介面對齊 openpyxl。

    `structure_detector.detect_structure` 只用到 title／max_row／max_column／
    cell()／merged_cells.ranges 這五樣，因此只需實作它們。
    """

    def __init__(
        self,
        title: str,
        grid: list[list[Any]],
        merged: list[_MergedRange] | None = None,
    ):
        self.title = title
        self._grid = grid
        self.max_row = len(grid)
        self.max_column = max((len(r) for r in grid), default=0)
        self.merged_cells = _MergedCells(merged or [])

    def cell(self, row: int, column: int) -> _Cell:
        """1-indexed，超出範圍回傳空值而非拋錯（與 openpyxl 行為一致）。"""
        if 1 <= row <= self.max_row:
            line = self._grid[row - 1]
            if 1 <= column <= len(line):
                return _Cell(line[column - 1])
        return _Cell(None)


class GridWorkbook:
    """以 sheet 名稱索引的工作簿，介面對齊 openpyxl。"""

    def __init__(self, sheets: dict[str, GridWorksheet]):
        self._sheets = sheets
        self.sheetnames = list(sheets)
        self.warnings: list[str] = []
        """讀取過程的降級訊息，例如合併儲存格取不到。由呼叫方併入輸出警告。"""

    def __getitem__(self, name: str) -> GridWorksheet:
        return self._sheets[name]

    def close(self) -> None:
        """介面對齊用；純記憶體結構沒有需要釋放的資源。"""


# --------------------------------------------------------------------------
# 各格式的讀取
# --------------------------------------------------------------------------

def _coerce(value: Any) -> Any:
    """
    把空字串正規化成 None，把數字字串轉成數值。

    CSV 的每一格都是字串，若不轉型，`structure_detector` 會判定整張表都是文字，
    永遠找不到「數值佔比過半」的資料列，整份檔案就解析失敗了。
    """
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # 去掉千分位再判斷；(1,234) 這種會計負數格式一併處理
        candidate = text.replace(",", "")
        negative = candidate.startswith("(") and candidate.endswith(")")
        if negative:
            candidate = candidate[1:-1]
        try:
            number = float(candidate)
        except ValueError:
            return text
        return -number if negative else number
    return value


def _read_xls(path: Path) -> GridWorkbook:
    """用 xlrd 讀舊版 .xls，含合併儲存格。"""
    try:
        import xlrd
    except ImportError as e:
        raise UnsupportedFormatError(
            f"讀取 {path.name} 需要 xlrd 套件（pip install xlrd）"
        ) from e

    # xlrd 只有在 formatting_info=True 時才會填 merged_cells。少了它，
    # 政府報表常見的多層合併表頭會被當成一堆空格，「亞洲地區」橫跨的欄位
    # 就拿不到上層語意——實測預設值會讀到 0 個合併儲存格。
    # 部分 .xls 含 xlrd 不認得的格式記錄會讓 formatting_info=True 拋錯，
    # 故退回不含格式的讀法：資料仍正確，只是失去合併資訊。
    try:
        book = xlrd.open_workbook(path, formatting_info=True)
        formatting = True
    except Exception:
        book = xlrd.open_workbook(path, formatting_info=False)
        formatting = False

    sheets: dict[str, GridWorksheet] = {}

    for name in book.sheet_names():
        sheet = book.sheet_by_name(name)
        grid = [
            [_coerce(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        # xlrd 的 merged_cells 是 0-indexed 且上界不含，需轉成 openpyxl 的
        # 1-indexed 含上界
        merged = [
            _MergedRange(rlo + 1, rhi, clo + 1, chi)
            for rlo, rhi, clo, chi in (sheet.merged_cells if formatting else [])
        ]
        sheets[name] = GridWorksheet(name, grid, merged)

    workbook = GridWorkbook(sheets)
    if not formatting:
        workbook.warnings.append(
            f"{path.name}：無法讀取合併儲存格資訊，多層表頭可能解析不完整"
        )
    return workbook


def _read_csv(path: Path) -> GridWorkbook:
    """
    讀 CSV。單一工作表、無合併儲存格。

    多重編碼嘗試是必要的：政府開放資料的 CSV 常是 cp950，用 utf-8 開會拋
    UnicodeDecodeError；若吞掉這個錯誤改用 errors='ignore'，中文欄名會變成
    亂碼，欄位比對就全錯了。
    """
    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                rows = [[_coerce(cell) for cell in line] for line in csv.reader(handle)]
            break
        except (UnicodeDecodeError, LookupError) as e:
            last_error = e
    else:
        raise UnsupportedFormatError(
            f"{path.name} 無法以 {CSV_ENCODINGS} 任一編碼讀取：{last_error}"
        )

    return GridWorkbook({path.stem: GridWorksheet(path.stem, rows)})


def open_workbook(path: str | Path):
    """
    依副檔名開啟工作簿，回傳的物件介面一致（sheetnames / [name] / close）。

    .xlsx 直接交給 openpyxl——它原生支援合併儲存格與儲存格座標，
    轉成中間格式反而會流失資訊。
    """
    target = Path(path)
    suffix = target.suffix.lower()

    if suffix in XLSX_SUFFIXES:
        return openpyxl.load_workbook(target, data_only=True)
    if suffix in XLS_SUFFIXES:
        return _read_xls(target)
    if suffix in CSV_SUFFIXES:
        return _read_csv(target)

    raise UnsupportedFormatError(
        f"{target.name} 的格式 '{suffix}' 不支援；"
        f"目前支援 {sorted(SUPPORTED_SUFFIXES)}"
    )


def iter_workbooks(data_dir: str | Path) -> Iterator[tuple[Path, Any, list[str]]]:
    """
    走訪目錄下所有可解析的檔案，yield (路徑, 工作簿, 警告)。

    單一檔案損毀不中斷整批——決賽現場 11 份裡有一份開不起來，
    不該讓另外 10 份也算不出來。該檔的錯誤會放進警告讓人看見。
    """
    scan = scan_inputs(data_dir)
    if not scan.supported:
        raise FileNotFoundError(
            f"{data_dir} 底下沒有可解析的資料檔"
            f"（支援 {sorted(SUPPORTED_SUFFIXES)}）"
            + (f"；另有 {len(scan.unsupported)} 個不支援的檔案："
               f"{'、'.join(p.name for p in scan.unsupported)}"
               if scan.unsupported else "")
        )

    pending = list(scan.warnings)
    for path in scan.supported:
        try:
            workbook = open_workbook(path)
        except Exception as e:
            pending.append(f"{path.name} 開啟失敗（已略過）：{type(e).__name__}: {e}")
            continue
        pending.extend(getattr(workbook, "warnings", []))
        yield path, workbook, pending
        pending = []