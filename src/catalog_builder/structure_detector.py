"""
結構偵測：動態找出標題列／多層表頭／資料起始列，並把多層表頭合併成完整欄名。

不可硬編碼「第 1 列是欄名」——Step 1 實測 11 份政府統計報表，資料起始列
分佈在第 3～5 列（3:5檔 / 4:4檔 / 5:2檔），寫死任何一個列號都會錯 6/11。
掃描結果見 Task1/recon/structure_report.md。

**表頭層數不能從合併儲存格數量推斷**：表1-3 只有 1 個合併儲存格，卻是 3 層
表頭（洲別 → 國家 → 英文名）。唯一可靠的訊號是「數值佔比」——實測表頭列
一律 0.0、資料列 0.978，中間沒有灰色地帶。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .cell_tracker import to_a1

# --- 門檻常數：全部取自 Step 1 實測，非猜測 ---

NUMERIC_RATIO_THRESHOLD = 0.5
"""某列非空儲存格中，數值佔比超過此值即視為資料列。

實測分離度極高（表頭 0.0 vs 資料 0.978），故此門檻不敏感，
在 0.2～0.9 之間取值都會得到相同結果。"""

TITLE_MAX_NON_EMPTY = 2
"""非空儲存格不超過此數的列視為「標題」而非表頭。

實測標題列只有 1 格（如「歷年來臺旅客按國籍分」橫跨整表），
而最稀疏的真表頭列有 5 格。

此規則只擋得住「未合併」的標題。已合併的標題（如 A1:H1）填平後會變成
整列同值，得靠 `_spans_full_width` 才抓得到。"""

TITLE_MIN_WIDTH_RATIO = 0.8
"""整列同值且涵蓋此比例以上的欄寬，視為標題。

用「是否從第 1 欄開始橫跨整表」來區分標題與分組表頭：
標題（「歷年來臺旅客按停留夜數分」A1:N1）從第 1 欄橫跨到底；
分組表頭（「停留夜數 Length of Stay」B2:N2）讓出第 1 欄給維度欄（年度），
不會從第 1 欄開始。這個差異才是兩者的本質區別，不是格數多寡。"""

MAX_HEADER_SCAN = 20
"""最多往下掃幾列找資料起點。實測最深只到第 5 列，20 列已是極寬鬆的保險。"""

HEADER_JOIN = " > "
"""多層表頭的階層分隔符，如「亞洲地區 > 日本 Japan」。"""


def _is_number(v: Any) -> bool:
    """bool 是 int 的子類，但 Excel 的 TRUE/FALSE 不是統計數值，須排除。"""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


@dataclass
class ColumnInfo:
    """單一欄位的表頭資訊。"""

    index: int
    """1-indexed 欄號。"""

    letter: str
    """欄字母，如 'C'。"""

    layers: list[str]
    """由上而下的各層表頭文字，已去除空值與重複。"""

    @property
    def full_name(self) -> str:
        """合併後的完整欄名，如「亞洲地區 > 日本 Japan」。"""
        return HEADER_JOIN.join(self.layers)

    @property
    def leaf_name(self) -> str:
        """最下層表頭，通常是最具體的名稱（如「日本 Japan」）。"""
        return self.layers[-1] if self.layers else ""

    @property
    def is_empty(self) -> bool:
        """完全沒有表頭文字的欄位（多為分隔用空欄）。"""
        return not self.layers


@dataclass
class SheetStructure:
    """一個工作表的結構偵測結果。"""

    sheet_name: str
    title_rows: list[int] = field(default_factory=list)
    header_rows: list[int] = field(default_factory=list)
    data_start_row: int | None = None
    data_end_row: int | None = None
    max_column: int = 0
    columns: list[ColumnInfo] = field(default_factory=list)
    grid: list[list[Any]] = field(default_factory=list)
    """合併儲存格已填平的完整值矩陣，grid[r-1][c-1] 對應儲存格 (r, c)。"""

    warnings: list[str] = field(default_factory=list)

    @property
    def is_parsable(self) -> bool:
        """有找到資料區才算解析成功。"""
        return self.data_start_row is not None and bool(self.header_rows)

    def cell(self, row: int, col: int) -> Any:
        """取得填平後的儲存格值（1-indexed）。"""
        if 1 <= row <= len(self.grid) and 1 <= col <= self.max_column:
            return self.grid[row - 1][col - 1]
        return None

    def column_by_index(self, index: int) -> ColumnInfo | None:
        return next((c for c in self.columns if c.index == index), None)


def build_grid(ws: Worksheet) -> list[list[Any]]:
    """
    讀出完整值矩陣，並把合併儲存格「填平」。

    Excel 的合併儲存格只有左上角那格有值，其餘是 None。若不填平，
    「亞洲地區」橫跨 G:N 就只有 G 欄拿得到洲別，H～N 欄會失去上層語意。
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    for merged in ws.merged_cells.ranges:
        anchor = grid[merged.min_row - 1][merged.min_col - 1]
        if anchor is None:
            continue
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                grid[r - 1][c - 1] = anchor

    return grid


def _row_numeric_ratio(row: list[Any]) -> tuple[int, float]:
    """回傳 (非空格數, 數值佔非空格比例)。分母用非空格數，空格不該稀釋判斷。"""
    non_empty = [v for v in row if v is not None and str(v).strip() != ""]
    if not non_empty:
        return 0, 0.0
    numbers = sum(1 for v in non_empty if _is_number(v))
    return len(non_empty), numbers / len(non_empty)


def _find_data_start(grid: list[list[Any]]) -> int | None:
    """由上往下掃，第一個數值佔比過半的列即資料起點。"""
    for idx, row in enumerate(grid[:MAX_HEADER_SCAN], start=1):
        non_empty, ratio = _row_numeric_ratio(row)
        if non_empty and ratio > NUMERIC_RATIO_THRESHOLD:
            return idx
    return None


def _find_data_end(grid: list[list[Any]], data_start: int) -> int:
    """
    由下往上找最後一個有資料的列。

    政府報表常在資料區下方接註腳（「資料來源：交通部觀光署」等純文字列），
    若不裁掉會被當成資料列讀進來。
    """
    for idx in range(len(grid), data_start - 1, -1):
        non_empty, ratio = _row_numeric_ratio(grid[idx - 1])
        if non_empty and ratio > NUMERIC_RATIO_THRESHOLD:
            return idx
    return data_start


def _spans_full_width(row: list[Any], max_col: int) -> bool:
    """
    整列是同一個值，且從第 1 欄橫跨到（接近）最後一欄。

    這是合併標題填平後的特徵。分組表頭（「輪船 Sea」橫跨 C:D）雖然也是
    同值連續，但一定讓出第 1 欄給維度欄，不會從第 1 欄開始。
    """
    values = [v for v in row if v is not None and str(v).strip() != ""]
    if len(values) < 2 or len({str(v).strip() for v in values}) != 1:
        return False
    if row[0] is None or str(row[0]).strip() == "":
        return False
    return len(values) / max_col >= TITLE_MIN_WIDTH_RATIO


def _split_title_and_header(
    grid: list[list[Any]], data_start: int, max_col: int
) -> tuple[list[int], list[int]]:
    """
    把資料列以上的列分成「標題」與「表頭」。

    標題（如「歷年來臺旅客按國籍分」）不對應任何欄位，混進表頭會讓每個
    欄名都被冠上同一串表名，既冗長又對欄位比對毫無幫助。
    """
    above = list(range(1, data_start))
    if not above:
        return [], []

    titles, headers = [], []
    for r in above:
        row = grid[r - 1]
        non_empty, _ = _row_numeric_ratio(row)
        # 稀疏列要從第 1 欄開始才算標題。分組表頭（「亞洲地區」橫跨 B:C）
        # 一定讓出第 1 欄給維度欄，只有標題會從最左邊開始——少了這個條件，
        # 只跨兩欄的分組表頭會被誤判成標題而整個丟掉。
        starts_at_first = row and row[0] is not None and str(row[0]).strip() != ""
        is_title = (
            (non_empty <= TITLE_MAX_NON_EMPTY and starts_at_first)
            or _spans_full_width(row, max_col)
        )
        (titles if is_title else headers).append(r)

    # 保險：若全被判為標題（例如單欄極簡表），把最靠近資料的那列還原成表頭，
    # 否則會得到「零個表頭」而整張表無法命名欄位。
    if not headers:
        headers = [titles.pop()]

    return titles, headers


def _build_columns(
    grid: list[list[Any]], header_rows: list[int], max_col: int
) -> list[ColumnInfo]:
    """把各層表頭由上而下疊成完整欄名。"""
    columns = []
    for c in range(1, max_col + 1):
        layers: list[str] = []
        for r in header_rows:
            v = grid[r - 1][c - 1]
            if v is None:
                continue
            text = str(v).strip()
            # 換行在政府報表的表頭很常見（「年度\nYear」），一律壓成單行
            text = " ".join(text.split())
            # 合併儲存格填平後，同一格值會在多層重複出現，去掉才不會變
            # 「亞洲地區 > 亞洲地區 > 日本」
            if text and text not in layers:
                layers.append(text)
        columns.append(ColumnInfo(index=c, letter=to_a1(1, c)[:-1], layers=layers))
    return columns


def detect_structure(ws: Worksheet) -> SheetStructure:
    """
    偵測單一工作表的結構。

    回傳的 SheetStructure 帶 `grid`（合併已填平），後續 normalizer 直接用它，
    不必重讀一次 Excel。
    """
    grid = build_grid(ws)
    max_col = ws.max_column or 0
    structure = SheetStructure(sheet_name=ws.title, max_column=max_col, grid=grid)

    if not grid or not max_col:
        structure.warnings.append("工作表為空，無法解析")
        return structure

    data_start = _find_data_start(grid)
    if data_start is None:
        structure.warnings.append(
            f"前 {MAX_HEADER_SCAN} 列找不到數值佔比 > {NUMERIC_RATIO_THRESHOLD} 的列，"
            "可能整張表皆為文字或格式異常"
        )
        return structure

    if data_start == 1:
        structure.warnings.append("資料從第 1 列開始，此表沒有表頭，欄位將無法命名")
        structure.data_start_row = 1
        structure.data_end_row = _find_data_end(grid, 1)
        return structure

    titles, headers = _split_title_and_header(grid, data_start, max_col)
    structure.title_rows = titles
    structure.header_rows = headers
    structure.data_start_row = data_start
    structure.data_end_row = _find_data_end(grid, data_start)
    structure.columns = _build_columns(grid, headers, max_col)

    unnamed = [c.index for c in structure.columns if c.is_empty]
    if unnamed:
        structure.warnings.append(
            f"{len(unnamed)} 個欄位無表頭文字（欄號 {unnamed[:8]}），將略過不建卡"
        )

    return structure
