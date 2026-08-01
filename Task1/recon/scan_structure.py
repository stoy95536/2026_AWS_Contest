"""
Step 1：結構偵察（唯讀）

目的不是「認識這 11 份檔案」，而是**測量同一族群統計報表的結構變異範圍**，
用來校準 catalog_builder 的偵測門檻——TASK1.md 1a-2 節寫的「數值佔比 > 50%」
目前只是假設，沒有任何真實資料支撐。門檻要從實測分佈訂，不是用猜的。

本腳本純唯讀：不修改任何 Excel，不寫入 outputs/，只產出偵察報告。

輸出：
  Task1/recon/raw_scan.json        — 機器可讀的完整掃描結果
  Task1/recon/structure_report.md  — 人可讀的結構報告

用法：
  python Task1/recon/scan_structure.py [data_dir]
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# 掃描深度：前 N 列用來判斷表頭結構
SCAN_ROWS = 15

# 每個 sheet 取樣多少列來統計數值佔比（涵蓋資料區）
PROFILE_ROWS = 40

# 民國/西元年字串樣本：「53年1964」「106 年 2017」「2017」「民國106年」
_YEAR_PATTERN = re.compile(r"(\d{2,4})\s*年|\b(19|20)\d{2}\b")


def _is_number(v: Any) -> bool:
    """儲存格是否為可運算的數值（bool 不算，Excel 裡 TRUE/FALSE 不是統計值）。"""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _cell_kind(v: Any) -> str:
    """把儲存格分類，用於統計每列的組成。"""
    if v is None:
        return "empty"
    if _is_number(v):
        return "number"
    if isinstance(v, (datetime, date)):
        return "date"
    return "text"


def _truncate(v: Any, limit: int = 24) -> Any:
    """報告用：長字串截短，避免表格爆版。"""
    if isinstance(v, str) and len(v) > limit:
        return v[: limit - 1] + "…"
    return v


def profile_row(values: list[Any]) -> dict:
    """
    分析單一列的組成。

    numeric_ratio 是 header 結束列偵測的核心訊號：表頭列幾乎全是文字，
    資料列幾乎全是數字。但「幾乎」是多少，要看實測分佈才知道。
    """
    kinds = Counter(_cell_kind(v) for v in values)
    non_empty = len(values) - kinds["empty"]
    return {
        "non_empty": non_empty,
        "number": kinds["number"],
        "text": kinds["text"],
        "date": kinds["date"],
        # 分母用非空格數，空格不該稀釋判斷
        "numeric_ratio": round(kinds["number"] / non_empty, 3) if non_empty else 0.0,
    }


def scan_sheet(ws) -> dict:
    """掃描單一工作表的結構特徵。"""
    max_row = min(ws.max_row or 0, PROFILE_ROWS)
    max_col = ws.max_column or 0

    # --- 前 N 列原始值（含 None，None 的位置正是合併儲存格的破洞）---
    head_rows = []
    for r in range(1, min(max_row, SCAN_ROWS) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        head_rows.append({
            "row": r,
            "profile": profile_row(values),
            "values": [_truncate(v) for v in values[:12]],  # 只留前 12 欄，夠看結構
        })

    # --- 全域數值佔比剖面：用來找資料起始列 ---
    row_profiles = []
    for r in range(1, max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        p = profile_row(values)
        p["row"] = r
        row_profiles.append(p)

    # --- 第一個「數值佔比過半」的列，即現行規則會判定的資料起點 ---
    first_majority_numeric = next(
        (p["row"] for p in row_profiles if p["numeric_ratio"] > 0.5), None
    )

    # --- 合併儲存格：多層表頭的直接證據 ---
    merged = [str(m) for m in ws.merged_cells.ranges]

    # --- 年度字串樣本：掃前兩欄找像年份的值 ---
    year_samples = []
    for r in range(1, max_row + 1):
        for c in range(1, min(max_col, 3) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            if _YEAR_PATTERN.search(s):
                year_samples.append({"cell": f"{get_column_letter(c)}{r}", "value": s})
                break
        if len(year_samples) >= 8:
            break

    # --- 負數位置：TASK1.md 點名的「小計調整／缺值標記」風險 ---
    negatives = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if _is_number(v) and v < 0:
                negatives.append({"cell": f"{get_column_letter(c)}{r}", "value": v})
            if len(negatives) >= 10:
                break
        if len(negatives) >= 10:
            break

    return {
        "sheet_name": ws.title,
        "dimensions": {
            "max_row": ws.max_row,
            "max_column": max_col,
            "scanned_rows": max_row,
        },
        "merged_cell_count": len(merged),
        "merged_cells": merged[:20],
        "first_majority_numeric_row": first_majority_numeric,
        "head_rows": head_rows,
        "row_profiles": row_profiles,
        "year_samples": year_samples,
        "negative_count_in_scan": len(negatives),
        "negatives": negatives,
    }


def scan_file(path: Path) -> dict:
    """掃描單一 Excel 檔案的所有工作表。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        return {
            "file_name": path.name,
            "sheet_count": len(wb.sheetnames),
            "sheets": [scan_sheet(wb[name]) for name in wb.sheetnames],
        }
    finally:
        wb.close()


def render_report(scan: list[dict]) -> str:
    """把掃描結果寫成人可讀的 Markdown，重點放在「變異範圍」。"""
    lines: list[str] = []
    w = lines.append

    w("# Step 1 結構偵察報告")
    w("")
    w(f"> 產生時間：{datetime.now():%Y-%m-%d %H:%M}　|　掃描檔案數：{len(scan)}")
    w("")
    w("本報告用途：測量統計報表的結構變異範圍，作為 `catalog_builder` 偵測門檻的依據。")
    w("**這些檔案是開發期測試集，非決賽檔案**——因此結論要看「變異範圍」，不是個別檔案的值。")
    w("")

    # ---------- 總覽 ----------
    w("## 1. 總覽")
    w("")
    w("| 檔案 | 工作表 | 列 x 欄 | 合併儲存格 | 首個數值過半列 | 掃描區負數 |")
    w("|---|---|---|---|---|---|")
    for f in scan:
        for s in f["sheets"]:
            d = s["dimensions"]
            w(
                f"| {f['file_name']} | {s['sheet_name']} | "
                f"{d['max_row']} x {d['max_column']} | {s['merged_cell_count']} | "
                f"{s['first_majority_numeric_row'] or '—'} | {s['negative_count_in_scan']} |"
            )
    w("")

    # ---------- 變異範圍：門檻校準的關鍵 ----------
    starts = [
        s["first_majority_numeric_row"]
        for f in scan
        for s in f["sheets"]
        if s["first_majority_numeric_row"]
    ]
    merges = [s["merged_cell_count"] for f in scan for s in f["sheets"]]
    w("## 2. 結構變異範圍（訂門檻的依據）")
    w("")
    if starts:
        w(f"- **資料起始列**：最小 {min(starts)}、最大 {max(starts)}、"
          f"分佈 {sorted(Counter(starts).items())}")
        w(f"  → 若寫死「第 N 列是欄名」，會錯 {len([x for x in starts if x != starts[0]])} / {len(starts)} 個工作表")
    if merges:
        w(f"- **合併儲存格數量**：最小 {min(merges)}、最大 {max(merges)}")
        w(f"  → {sum(1 for m in merges if m > 0)} / {len(merges)} 個工作表有合併儲存格（多層表頭證據）")
    w("")

    # ---------- 逐表細節 ----------
    w("## 3. 逐工作表細節")
    w("")
    for f in scan:
        w(f"### 📄 {f['file_name']}")
        w("")
        for s in f["sheets"]:
            w(f"#### 工作表：`{s['sheet_name']}`")
            w("")
            w(f"- 尺寸：{s['dimensions']['max_row']} 列 x {s['dimensions']['max_column']} 欄")
            w(f"- 合併儲存格：{s['merged_cell_count']} 個"
              + (f"，前幾個：`{'`, `'.join(s['merged_cells'][:6])}`" if s["merged_cells"] else ""))
            w(f"- 首個數值佔比 > 50% 的列：**{s['first_majority_numeric_row'] or '無'}**")
            w("")

            w("**前 12 列數值佔比剖面**（找表頭/資料分界）")
            w("")
            w("| 列 | 非空 | 數字 | 文字 | 數值佔比 | 前幾格內容 |")
            w("|---|---|---|---|---|---|")
            for hr in s["head_rows"][:12]:
                p = hr["profile"]
                preview = " │ ".join(
                    "" if v is None else str(v) for v in hr["values"][:6]
                )
                w(f"| {hr['row']} | {p['non_empty']} | {p['number']} | {p['text']} | "
                  f"{p['numeric_ratio']} | {preview} |")
            w("")

            if s["year_samples"]:
                w("**年度字串樣本**（正規化難度）")
                w("")
                for ys in s["year_samples"][:6]:
                    w(f"- `{ys['cell']}` = `{ys['value']}`")
                w("")

            if s["negatives"]:
                w("**負數位置**（疑似小計調整／缺值標記，不可默默加總）")
                w("")
                for ng in s["negatives"][:6]:
                    w(f"- `{ng['cell']}` = `{ng['value']}`")
                w("")
    return "\n".join(lines)


def main(data_dir: str) -> int:
    src = Path(data_dir)
    if not src.is_dir():
        print(f"[ERROR] 找不到資料目錄: {src}")
        return 1

    files = sorted(src.glob("*.xlsx"))
    if not files:
        print(f"[ERROR] {src} 底下沒有 .xlsx")
        return 1

    print(f"[INFO] 掃描 {len(files)} 份檔案（唯讀）")
    scan = []
    for p in files:
        print(f"  - {p.name}")
        try:
            scan.append(scan_file(p))
        except Exception as e:
            print(f"    [WARN] 解析失敗: {type(e).__name__}: {e}")
            scan.append({"file_name": p.name, "error": f"{type(e).__name__}: {e}", "sheets": []})

    out_dir = Path(__file__).parent
    json_path = out_dir / "raw_scan.json"
    md_path = out_dir / "structure_report.md"

    json_path.write_text(
        json.dumps(scan, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    md_path.write_text(render_report(scan), encoding="utf-8")

    print(f"[OK] {json_path}")
    print(f"[OK] {md_path}")
    return 0


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).parents[1] / "data")
    sys.exit(main(target))